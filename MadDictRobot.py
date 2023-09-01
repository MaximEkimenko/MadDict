from pathlib import Path
from selenium import webdriver
import time
import os
import openpyxl
from selenium.webdriver.common.by import By
import wave
import pyaudio
from pydub import AudioSegment, silence

outputFileName = 'MadDictBase.xlsx'
# создание листа существующих имен
# открытие существующего словаря для проверки содержимого
exist_wb = openpyxl.load_workbook(outputFileName, read_only=True)
exist_sh = exist_wb['Sheet']
existingList = []
for val in exist_sh.values:
    existingList.append(val[0])  # лист существующих имен
wb_input = openpyxl.load_workbook('gtlist.xlsx')  # файл сохранённых переводов с google translate
sh_input = wb_input['Сохраненные переводы']
word_list = []  # список английских слов
tr_list = []  # список переводов
words = dict()
for val in sh_input.values:  # заполнение списков
    words.update({val[2]: val[3]})
    word_list.append(val[2])
    tr_list.append(val[3])
path = os.getcwd() + r"\audio" + "\\"  # путь сохранения аудио файлов
# результирующий файл Excel
res_wb = openpyxl.Workbook()
res_sh = res_wb.active

for row in exist_sh.values:  # копирование значений в рез лист если они уже были записаны ранее
    res_sh.append(row)

for word in word_list:  # запуск хрома только если есть новое слово
    if word not in existingList:  # если слово ещё не добавлено
        driver = webdriver.Chrome()  # объект хрома selenium
        driver.get("https://translate.google.ru/?hl=ru&sl=en&tl=ru&op=translate")  # открытие страницы google translate
        break
# смещение курсора на величину существующего списка
if existingList:
    i = len(existingList)
else:
    i = 0
for word in word_list:
    if word not in existingList:  # если слово ещё не добавлено
        i = i + 1
        areaElems = driver.find_elements(By.TAG_NAME, 'textarea')  # поле ввода значений
        areaElems[0].click()  # активация кликом поля
        time.sleep(1)
        areaElems[0].clear()  # очистка поля
        time.sleep(1)  # пауза 1 сек
        areaElems[0].send_keys(word)  # ввод варианта перевода
        time.sleep(2)  # пауза 1 сек
        areaElems[0].click()  # активация кликом поля
        time.sleep(3)  # пауза 1 сек
        tVars = []  # список вариантов перевода
        tElems = driver.find_elements(By.CLASS_NAME, 'kgnlhe')  # элементы перевода
        for elem in tElems:
            if elem.text != '':
                tVars.append(elem.text)
        voiceBtn = driver.find_element(By.XPATH,  # кнопка прослушать
                                       '//*[@id="yDmH0d"]/c-wiz/div/div[2]/c-wiz/div[2]/c-wiz/div[1]/div[2]/div['
                                       '3]/c-wiz[1]/div[4]/div[2]/div/div['
                                       '2]/span/button/div[3]')
        time.sleep(1)  # пауза 1 сек
        voiceBtn.click()  # нажатие кнопки прослушать
        # запись произношения в аудио файл
        chunk = 512  # Запись кусками по битам сэмпла
        sample_format = pyaudio.paInt16  # 16 бит на выборку
        channels = 1  # моно канал
        rate = 16000  # Запись со скоростью 44100 выборок(samples) в секунду
        seconds = 5  # 5 секунд будет достаточно
        filename = str(Path(path, word + '.wav'))  # полный путь до файла
        p = pyaudio.PyAudio()  # создать интерфейс для PortAudio
        print('Recording...', filename)
        stream = p.open(format=sample_format,  # создание потока
                        channels=channels,
                        rate=rate,
                        frames_per_buffer=chunk,
                        input_device_index=1,  # индекс устройства с которого будет идти запись звука
                        input=True)
        frames = []  # инициализировать массив для хранения кадров
        # Хранить данные в блоках в течение seconds секунд
        for sti in range(0, int(rate / chunk * seconds)):
            data = stream.read(chunk)
            frames.append(data)
        # Остановить и закрыть поток
        stream.stop_stream()
        stream.close()
        # Завершить интерфейс PortAudio
        p.terminate()
        print('Finished recording! ', filename)
        # Сохранить записанные данные в виде файла WAV
        wf = wave.open(filename, 'wb')
        wf.setnchannels(channels)
        wf.setsampwidth(p.get_sample_size(sample_format))
        wf.setframerate(rate)
        wf.writeframes(b''.join(frames))
        wf.close()
        # обрезание пустоты в файле записи (сокращение файла)
        sound = AudioSegment.from_file(filename, format="wav")  # открытие только что записанного файла
        timeList = silence.detect_nonsilent(sound, silence_thresh=-200,
                                            seek_step=5)  # получение НЕ пустот в начало, конец в мс
        # сохранение в тот же файл среза без пустот
        sliceRez = sound[timeList[0][0]:timeList[0][1] + 500]
        sliceRez.export(filename, format="wav")
        # rez_list = []
        # сохранение вариантов перевода со словом
        # rez_list.append(word)
        res_sh.cell(row=i, column=1).value = word
        if tVars:
            # rez_list.append(trList[i-1]+ "," + ",".join(tVars))
            res_sh.cell(row=i, column=2).value = words[word] + "," + ",".join(tVars)
        else:
            #   rez_list.append(trList[i-1])
            # res_sh.append(rez_list)
            res_sh.cell(row=i, column=2).value = words[word]
        res_sh.cell(row=i, column=3).value = 6  # заполнение нулевых значений для количества повторов
        res_sh.cell(row=i, column=4).value = 0  # заполнение нулевых значений для количества выполненных повторов
        res_wb.save(outputFileName)  # сохранение файла с результатами
    else:
        print(word, ' is already added!')

# res_wb.save(outputFileName)  # сохранение файла с результатами

