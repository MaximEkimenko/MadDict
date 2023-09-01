import openpyxl
import random
from typing import Any


def get_random_line(exel_file) -> dict and int and int:
    """
    Читает excel базу в словарь
    :param exel_file: путь к базе слов в формате excel
    :return: словарь случайной строки из базы вида:
    {word: (строка переводов, общее количество повторов, количество сделанных повторов)}
    """
    ex_wb = openpyxl.load_workbook(exel_file, read_only=True)
    ex_sh = ex_wb.active
    result_dict = dict()
    i = 0  # инкремент словаря
    g = random.randint(1, ex_sh.max_row)
    words_studied = set()
    random_line = ''
    for row in ex_sh.iter_rows(min_row=1, max_row=ex_sh.max_row, min_col=1, max_col=ex_sh.max_column,
                               values_only=True):
        i += 1
        result_dict.update({row[0]: (row[1], row[2], row[3])})
        if i == g and 0 < row[2] <= 6:
            random_line = (row[0], row[1], row[2], row[3])
            break
        else:
            random_line = 'Всё изучено'
        if row[3] >= 6:
            words_studied.add(row[3])

    return random_line, i, ex_sh.max_row, len(words_studied)


def update_line(excel_file: str, col_letter: str, base_row: int, value: Any) -> None:
    """
    Обновляет строку base_row в excel_file
    :param excel_file: - путь к файлу Excel
    :param col_letter: - литера столбца для редактирования
    :param base_row: - номер корректируемого ряда
    :param value: новое значение ячейки
    :return: None
    """
    ex_wb = openpyxl.load_workbook(excel_file)
    ex_sh = ex_wb.active
    ex_sh[col_letter+str(base_row)].value = value
    ex_wb.save(excel_file)
    ex_wb.close()
    return None


if __name__ == '__main__':
    ex_file = r'D:\АСУП\Python\Projects\MadDict\MadDictBase.xlsx'
    # print(get_random_line(ex_file))
    # update_line(excel_file=ex_file, col_letter='E', base_row=1, value=15000)
