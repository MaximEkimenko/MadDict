import os

import openpyxl

outputFileName = 'MadDictBase.xlsx'
# создание листа существующих имен
# открытие существующего словаря для проверки содержимого
exist_wb = openpyxl.load_workbook(outputFileName, read_only=True)
exist_sh = exist_wb['Sheet']
existingList = []
for val in exist_sh.values:
    existingList.append(val[0])  # лист существующих имен
wb_input = openpyxl.load_workbook('gtlist.xlsx')  # файл сохранённых переводов с google translate
sh_input = wb_input['Сохраненные переводы']
word_list = []  # список английских слов
tr_list = []  # список переводов
words = dict()
for val in sh_input.values:  # заполнение списков
    words.update({val[2]: val[3]})
    word_list.append(val[2])
    tr_list.append(val[3])
path = os.getcwd() + r"\audio" + "\\"  # путь сохранения аудио файлов
print(words)
# результирующий файл Excel
res_wb = openpyxl.Workbook()